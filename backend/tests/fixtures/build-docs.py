"""Deterministic synthetic originals, not edits or validation attestations."""
import datetime
import hashlib
import io
import json
from pathlib import Path
import sys
import zipfile

from lxml import etree
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook.properties import CalcProperties
from PIL import Image, ImageDraw, ImageFont
import reportlab
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

STAMP = datetime.datetime(2026, 1, 1)
NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
CT = 'http://schemas.openxmlformats.org/package/2006/content-types'
REL = 'http://schemas.openxmlformats.org/package/2006/relationships'
FONT_PATH = Path(reportlab.__file__).parent / 'fonts/Vera.ttf'


def stable_zip(parts):
    result = io.BytesIO()
    with zipfile.ZipFile(result, 'w', compression=zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
        for name in sorted(parts):
            entry = zipfile.ZipInfo(name, (2026, 1, 1, 0, 0, 0))
            entry.compress_type = zipfile.ZIP_DEFLATED
            entry.external_attr = 0o600 << 16
            archive.writestr(entry, parts[name])
    return result.getvalue()


def spreadsheet():
    workbook = Workbook()
    workbook.properties.creator = 'SiraGPT QA - SYNTHETIC'
    workbook.properties.created = STAMP
    workbook.properties.modified = STAMP
    workbook.calculation = CalcProperties(fullCalcOnLoad=True, forceFullCalc=True, calcMode='auto')
    sheet = workbook.active
    sheet.title = 'Presupuesto'
    summary = workbook.create_sheet('Resumen')
    sheet.merge_cells('A1:D1')
    sheet['A1'] = 'PRESUPUESTO SINTETICO - NO DATOS DE CLIENTES'
    sheet.append(['Entradas azules; formulas negras. Todos los importes son datos de prueba.'])
    sheet.append(['Concepto', 'Entrada', 'Unidades', 'Subtotal'])
    for index, (label, cost, units) in enumerate([('Equipo A', 100, 10), ('Equipo B', 120, 5), ('Equipo C', 80, 12)], start=4):
        sheet.append([label, cost, units, f'=B{index}*C{index}'])
        sheet.cell(index, 2).comment = Comment('Dato sintetico para pruebas; editar B4, B5 y B6 en G4.', 'QA')
        sheet.cell(index, 2).font = Font(name='Arial', color='0000FF')
    sheet['A8'] = 'Encabezado compartido'
    sheet['A9'] = 'Encabezado compartido'
    summary['A1'] = 'RESUMEN SINTETICO'
    summary['A3'] = 'Concepto'
    summary['B3'] = 'Total (S/)'
    summary['A4'] = 'Neto'
    summary['B4'] = '=SUM(Presupuesto!D4:D6)'
    summary['A5'] = 'IGV de prueba'
    summary['B5'] = '=B4*B8'
    summary['A6'] = 'Total'
    summary['B6'] = '=SUM(B4:B5)'
    summary['A8'] = 'Tasa de prueba'
    summary['B8'] = 0.18
    summary['B8'].comment = Comment('Supuesto sintetico, no asesoramiento tributario.', 'QA')
    summary['A10'] = 'Encabezado compartido'
    sheet.conditional_formatting.add('D4:D6', CellIsRule(operator='greaterThan', formula=['900'], fill=PatternFill('solid', fgColor='FFF2CC')))
    chart = BarChart()
    chart.title = 'Subtotales de prueba'
    chart.y_axis.title = 'S/'
    chart.x_axis.title = 'Equipo'
    chart.add_data(Reference(sheet, min_col=4, min_row=3, max_row=6), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=4, max_row=6))
    chart.height, chart.width = 8, 15
    summary.add_chart(chart, 'D3')
    for tab in workbook:
        tab.sheet_view.showGridLines = False
        tab.freeze_panes = 'B4'
        tab.sheet_properties.pageSetUpPr.fitToPage = True
        tab.page_setup.orientation = 'landscape'
        tab.page_setup.paperSize = tab.PAPERSIZE_A4
        tab.page_setup.fitToWidth = 1
        tab.page_setup.fitToHeight = 1
        for col in ['A', 'B', 'C', 'D']:
            tab.column_dimensions[col].width = 27 if col == 'A' else 19
        for row in tab:
            for cell in row:
                if cell.coordinate not in ('B4', 'B5', 'B6') or tab != sheet:
                    cell.font = Font(name='Arial', size=11, color='008000' if cell.data_type == 'f' and '!' in cell.value else '000000')
                cell.alignment = Alignment(vertical='center')
                if cell.row == 3 or cell.row == 1:
                    cell.font = Font(name='Arial', bold=True, color='FFFFFF')
                    cell.fill = PatternFill('solid', fgColor='174A45')
    sheet.print_options.horizontalCentered = True
    sheet.print_area = 'A1:D10'
    summary.print_area = 'A1:L19'
    raw = io.BytesIO()
    workbook.save(raw)
    with zipfile.ZipFile(raw) as archive:
        parts = {name: archive.read(name) for name in archive.namelist()}
    # openpyxl overwrites modified during save, even when properties.modified
    # was assigned. Normalize package metadata after generation, not at edit.
    core = etree.fromstring(parts['docProps/core.xml'])
    for key in ('created', 'modified'):
        core.find(f'{{http://purl.org/dc/terms/}}{key}').text = '2026-01-01T00:00:00Z'
    parts['docProps/core.xml'] = etree.tostring(core)
    # openpyxl writes inline strings. Convert them into one real shared-string
    # table so G5 can prove that editing one cell does not alter its peers.
    strings, count = [], 0
    cached = {'xl/worksheets/sheet1.xml': {'D4': '1000', 'D5': '600', 'D6': '960'},
              'xl/worksheets/sheet2.xml': {'B4': '2560', 'B5': '460.8', 'B6': '3020.8'}}
    for name in cached:
        root = etree.fromstring(parts[name])
        for cell in root.findall(f'.//{{{NS}}}c'):
            if cell.get('t') == 'inlineStr':
                text = ''.join(cell.find(f'{{{NS}}}is').itertext())
                if text not in strings:
                    strings.append(text)
                cell.remove(cell.find(f'{{{NS}}}is'))
                cell.set('t', 's')
                etree.SubElement(cell, f'{{{NS}}}v').text = str(strings.index(text))
                count += 1
            if cell.get('r') in cached[name]:
                cell.find(f'{{{NS}}}v').text = cached[name][cell.get('r')]
        parts[name] = etree.tostring(root)
    table = etree.Element(f'{{{NS}}}sst', nsmap={None: NS}, count=str(count), uniqueCount=str(len(strings)))
    for text in strings:
        etree.SubElement(etree.SubElement(table, f'{{{NS}}}si'), f'{{{NS}}}t').text = text
    parts['xl/sharedStrings.xml'] = etree.tostring(table)
    rels = etree.fromstring(parts['xl/_rels/workbook.xml.rels'])
    etree.SubElement(rels, f'{{{REL}}}Relationship', Id='rIdSyntheticSharedStrings',
                     Type='http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings', Target='sharedStrings.xml')
    parts['xl/_rels/workbook.xml.rels'] = etree.tostring(rels)
    types = etree.fromstring(parts['[Content_Types].xml'])
    etree.SubElement(types, f'{{{CT}}}Override', PartName='/xl/sharedStrings.xml',
                     ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml')
    parts['[Content_Types].xml'] = etree.tostring(types)
    return stable_zip(parts)


def build(output):
    output = Path(output)
    output.mkdir(parents=True, exist_ok=True)
    pdfmetrics.registerFont(TTFont('FixtureVera', str(FONT_PATH)))
    graphic = Image.new('RGB', (480, 260), '#174A45')
    draw = ImageDraw.Draw(graphic)
    font = ImageFont.truetype(str(FONT_PATH), 26)
    draw.text((32, 35), 'SIRA / PRUEBA SINTETICA', fill='white', font=font)
    for x, height in [(60, 50), (155, 85), (250, 120), (345, 150)]:
        draw.rectangle((x, 225-height, x+50, 225), fill='#91D3BB')
    graphic.save(output / 'fixture-diagram.png')
    (output / 'presupuesto.xlsx').write_bytes(spreadsheet())
    for name, pages in [('informe.pdf', 2), ('anexo.pdf', 1)]:
        document = canvas.Canvas(str(output / name), pagesize=(595, 842), invariant=1, pageCompression=1)
        document.setAuthor('SiraGPT QA - SYNTHETIC')
        document.setTitle(f'{name} - synthetic fixture')
        for page in range(1, pages+1):
            document.setFont('FixtureVera', 18)
            document.drawString(50, 775, f'INFORME SINTETICO {name} / {page}')
            document.setFont('FixtureVera', 12)
            document.drawString(50, 730, f'Contenido original verificable. Fecha: 2026-01-01. Pagina {page}.')
            document.drawString(50, 705, 'No contiene informacion de personas ni clientes reales.')
            if name == 'informe.pdf' and page == 1:
                document.drawString(50, 655, 'Campo de formulario:')
                document.acroForm.textfield(name='codigo_prueba', value='SINTETICO-001', x=50, y=610,
                                            width=220, height=25, borderWidth=1, forceBorder=True)
            document.showPage()
        document.save()
    scanned = Image.new('RGB', (1190, 1684), 'white')
    draw = ImageDraw.Draw(scanned)
    draw.text((100, 160), 'DOCUMENTO ESCANEADO SINTETICO', fill='black', font=ImageFont.truetype(str(FONT_PATH), 40))
    draw.text((100, 260), 'Este parrafo solo existe como imagen.', fill='black', font=ImageFont.truetype(str(FONT_PATH), 32))
    document = canvas.Canvas(str(output / 'escaneado.pdf'), pagesize=(595, 842), invariant=1)
    document.drawImage(ImageReader(scanned), 0, 0, width=595, height=842)
    document.showPage()
    document.save()
    (output / 'python-build.json').write_text(json.dumps({'font': 'Bitstream Vera (ReportLab bundled)',
      'fontSha256': hashlib.sha256(FONT_PATH.read_bytes()).hexdigest(), 'reportlab': reportlab.Version}, sort_keys=True)+'\n')


if __name__ == '__main__':
    if len(sys.argv) != 2:
        raise SystemExit('usage: build-docs.py ABSOLUTE_OUTPUT_DIRECTORY')
    if not Path(sys.argv[1]).is_absolute():
        raise SystemExit('output directory must be absolute')
    build(sys.argv[1])
