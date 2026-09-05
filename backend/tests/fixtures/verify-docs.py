"""Real-tool checks of generated originals. No editor or isolation attestation."""
import hashlib
import importlib.util
import json
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest
import zipfile

from lxml import etree
from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(sys.argv.pop(1)).resolve()
SCRIPT = Path(__file__).parents[2] / 'src/modules/doc-sandbox/validation/validator.py'
spec = importlib.util.spec_from_file_location('real_validator', SCRIPT)
validator = importlib.util.module_from_spec(spec)
spec.loader.exec_module(validator)
NS = {**validator.NS, 'p': 'http://schemas.openxmlformats.org/presentationml/2006/main',
      'rel': 'http://schemas.openxmlformats.org/package/2006/relationships'}


def parts(name):
    with zipfile.ZipFile(ROOT / name) as archive:
        return {name: archive.read(name) for name in archive.namelist()}


def xml(data):
    return etree.fromstring(data)


class ComplexOriginals(unittest.TestCase):
    def test_manifest_files_and_all_part_hashes(self):
        manifest = json.loads((ROOT / 'manifest.json').read_text())
        self.assertFalse(manifest['editorExecuted'])
        self.assertFalse(manifest['specificationGoldensSatisfied'])
        self.assertTrue(manifest['synthetic'])
        self.assertFalse(manifest['userDocuments'])
        self.assertEqual(len(manifest['files']), 6)
        for entry in manifest['files']:
            data = (ROOT / entry['name']).read_bytes()
            self.assertEqual(hashlib.sha256(data).hexdigest(), entry['sha256'])
            self.assertEqual(len(data), entry['bytes'])
            if 'parts' in entry:
                package = parts(entry['name'])
                self.assertEqual(set(package), set(entry['parts']))
                for name, expected in entry['parts'].items():
                    self.assertEqual(hashlib.sha256(package[name]).hexdigest(), expected)

    def test_office_packages_pass_real_independent_structural_inspector(self):
        for name in ['tesis.docx', 'presupuesto.xlsx', 'defensa.pptx']:
            with self.subTest(name=name):
                inventory = validator.inspect(ROOT / name, name)
                self.assertGreater(len(inventory['units']), 20)
                self.assertEqual(inventory['sha256'], hashlib.sha256((ROOT / name).read_bytes()).hexdigest())

    def test_word_three_runs_have_distinct_properties_and_exact_phrase(self):
        root = xml(parts('tesis.docx')['word/document.xml'])
        matching = [p for p in root.findall('.//w:p', NS)
                    if ''.join(p.xpath('.//w:t/text()', namespaces=NS)) == 'La gestion de compras reduce los costos operativos.']
        self.assertEqual(len(matching), 1)
        runs = matching[0].findall('w:r', NS)
        self.assertEqual(len(runs), 3)
        properties = [etree.tostring(run.find('w:rPr', NS)) for run in runs]
        self.assertEqual(len(set(properties)), 3)
        self.assertIsNotNone(runs[1].find('w:rPr/w:b', NS))
        self.assertIsNotNone(runs[2].find('w:rPr/w:i', NS))

    def test_word_headers_pagefield_footnotes_list_table_image_textbox(self):
        package = parts('tesis.docx')
        self.assertIn('word/header1.xml', package)
        self.assertIn(b'UNIVERSIDAD DE PRUEBA', package['word/header1.xml'])
        self.assertIn(b'PAGE', package['word/footer1.xml'])
        footnotes = xml(package['word/footnotes.xml'])
        for note_id in ['1', '2']:
            self.assertEqual(len(footnotes.xpath(f'//w:footnote[@w:id="{note_id}"]', namespaces=NS)), 1)
        root = xml(package['word/document.xml'])
        self.assertGreaterEqual(len(root.findall('.//w:numPr', NS)), 2)
        self.assertEqual(len(root.findall('.//w:tbl', NS)), 1)
        self.assertEqual(len(root.findall('.//w:drawing', NS)), 1)
        self.assertEqual(len(root.findall('.//w:txbxContent', NS)), 1)
        self.assertTrue(any(name.startswith('word/media/') for name in package))

    def test_xlsx_cross_sheet_formulas_chart_condition_merge(self):
        package = parts('presupuesto.xlsx')
        workbook = load_workbook(ROOT / 'presupuesto.xlsx', data_only=False)
        self.assertEqual(workbook.sheetnames, ['Presupuesto', 'Resumen'])
        self.assertEqual(workbook['Resumen']['B4'].value, '=SUM(Presupuesto!D4:D6)')
        self.assertEqual(workbook['Resumen']['B5'].value, '=B4*B8')
        self.assertEqual(workbook['Resumen']['B6'].value, '=SUM(B4:B5)')
        for index in [4, 5, 6]:
            self.assertEqual(workbook['Presupuesto'][f'D{index}'].value, f'=B{index}*C{index}')
        self.assertIn('A1:D1', workbook['Presupuesto'].merged_cells)
        self.assertEqual(len(workbook['Presupuesto'].conditional_formatting), 1)
        self.assertEqual(len(workbook['Resumen']._charts), 1)
        self.assertIn('xl/charts/chart1.xml', package)
        self.assertTrue(workbook.calculation.fullCalcOnLoad)
        workbook.close()

    def test_xlsx_shared_string_is_really_reused_by_three_cells(self):
        package = parts('presupuesto.xlsx')
        strings = xml(package['xl/sharedStrings.xml'])
        values = [''.join(item.itertext()) for item in strings.findall('s:si', NS)]
        shared_id = str(values.index('Encabezado compartido'))
        for name, coordinate in [('sheet1.xml', 'A8'), ('sheet1.xml', 'A9'), ('sheet2.xml', 'A10')]:
            cell = xml(package[f'xl/worksheets/{name}']).xpath(f'//s:c[@r="{coordinate}"]', namespaces=NS)[0]
            self.assertEqual(cell.get('t'), 's')
            self.assertEqual(cell.find('s:v', NS).text, shared_id)

    def test_xlsx_real_libreoffice_recalculation_matches_original_caches(self):
        original_hash = hashlib.sha256((ROOT / 'presupuesto.xlsx').read_bytes()).hexdigest()
        with tempfile.TemporaryDirectory(prefix='fixture-recalc-') as temporary:
            destination = Path(temporary) / 'recalc'
            self.assertEqual(validator.formula_errors(ROOT / 'presupuesto.xlsx', destination), set())
            recalculated = load_workbook(destination / 'presupuesto.xlsx', data_only=True)
            original = load_workbook(ROOT / 'presupuesto.xlsx', data_only=True)
            for cell, expected in [('B4', 2560), ('B5', 460.8), ('B6', 3020.8)]:
                self.assertAlmostEqual(recalculated['Resumen'][cell].value, expected)
                self.assertAlmostEqual(original['Resumen'][cell].value, expected)
            recalculated.close()
            original.close()
        self.assertEqual(hashlib.sha256((ROOT / 'presupuesto.xlsx').read_bytes()).hexdigest(), original_hash)

    def test_pptx_eight_slides_notes_image_and_distinct_layouts(self):
        package = parts('defensa.pptx')
        presentation = xml(package['ppt/presentation.xml'])
        self.assertEqual(len(presentation.findall('p:sldIdLst/p:sldId', NS)), 8)
        layouts = set()
        for index in range(1, 9):
            slide = xml(package[f'ppt/slides/slide{index}.xml'])
            self.assertGreaterEqual(len(slide.findall('.//p:pic', NS)), 1)
            notes = xml(package[f'ppt/notesSlides/notesSlide{index}.xml'])
            text = ''.join(notes.xpath('//a:t/text()', namespaces=NS))
            self.assertIn(f'diapositiva {index}', text)
            rels = xml(package[f'ppt/slides/_rels/slide{index}.xml.rels'])
            layouts.update(rel.get('Target') for rel in rels if rel.get('Type').endswith('/slideLayout'))
        self.assertEqual(len(layouts), 2)
        self.assertIn(b'Objetivos originales 2026', package['ppt/slides/slide3.xml'])

    def test_pdf_embedded_font_form_and_merge_page_counts(self):
        first, second = PdfReader(ROOT / 'informe.pdf', strict=True), PdfReader(ROOT / 'anexo.pdf', strict=True)
        self.assertEqual((len(first.pages), len(second.pages)), (2, 1))
        self.assertEqual(first.get_fields()['codigo_prueba']['/V'], 'SINTETICO-001')
        for reader in [first, second]:
            for page in reader.pages:
                fonts = page['/Resources']['/Font'].get_object().values()
                self.assertTrue(any('/FontFile2' in font.get_object().get('/FontDescriptor', {}) for font in fonts))
                self.assertIn('Contenido original verificable.', page.extract_text())

    def test_scanned_pdf_has_image_and_no_text_layer(self):
        reader = PdfReader(ROOT / 'escaneado.pdf', strict=True)
        self.assertEqual(len(reader.pages), 1)
        self.assertEqual(reader.pages[0].extract_text(), '')
        self.assertEqual(len(reader.pages[0].images), 1)
        rendered_text = subprocess.run(['pdftotext', str(ROOT / 'escaneado.pdf'), '-'], capture_output=True, check=True).stdout
        self.assertEqual(rendered_text.strip(), b'')

    def test_real_office_render_has_expected_page_counts_and_text(self):
        with tempfile.TemporaryDirectory(prefix='complex-fixture-render-') as temporary:
            for name, count in [('tesis.docx', 2), ('presupuesto.xlsx', 2), ('defensa.pptx', 8)]:
                with self.subTest(name=name):
                    output = validator.convert_pdf(ROOT / name, Path(temporary) / name, name.split('.')[-1])
                    self.assertEqual(validator.pdf_pages(output), count)
                    text = subprocess.run(['pdftotext', str(output), '-'], capture_output=True, check=True).stdout.decode()
                    self.assertTrue(text.strip())
                    if name == 'tesis.docx':
                        self.assertIn('La gestion de compras reduce los costos operativos.', text.replace('\n', ' '))
                        self.assertIn('Cuadro de texto:', text)
                    if name == 'defensa.pptx':
                        self.assertIn('Objetivos originales 2026', text)


if __name__ == '__main__':
    unittest.main(verbosity=2)
