"""sgpt_xlsx — corporate / academic Excel helpers.

Uses openpyxl for richer formatting (conditional formatting, data
validation, charts) while exposing a simpler API than openpyxl itself.
Writes workbooks with:
  · Frozen header row
  · Header row styled (bold white on navy)
  · Auto-width columns
  · Alternating-row fill (zebra)
  · Number format per column kind
  · Optional data-validation dropdowns
  · Multiple sheets wiring (Raw Data, Descriptives, Reliability,
    Correlations)

Analysis helpers:
  · cronbach_alpha(arr)        — array of arrays of Likert scores
  · spearman_matrix(df_like)   — dict[name] → list, returns pairwise rho
  · descriptives(arr, label)   — mean/sd/median/min/max
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import statistics
import math


# ─── Colour palette ───────────────────────────────────────────────────────

NAVY = '1F3A68'
NAVY_LIGHT = '4A6FA5'
CREAM = 'F7F3EA'
ACCENT = 'D97706'
GOOD = '10B981'
BAD = 'EF4444'
WARN = 'F59E0B'
GREY_LIGHT = 'F3F4F6'
GREY_MED = 'E5E7EB'


def corporate_workbook():
    """Return an empty workbook with the default sheet removed."""
    wb = Workbook()
    # Keep first sheet, callers rename it.
    return wb


# ─── Table writer ────────────────────────────────────────────────────────

def write_table(
    ws, *, headers, rows,
    title=None, start_row=1, start_col=1,
    freeze_header=True, alt_rows=True,
    number_formats=None,  # dict col_name -> format string
    column_widths=None,   # dict col_name -> int width (chars)
    autofit=True,
):
    """Write a structured table with header styling + zebra + auto-width.

    Returns (header_row_index, last_row_index, last_col_index).
    """
    r = start_row
    c0 = start_col

    if title:
        ws.cell(row=r, column=c0, value=title).font = Font(bold=True, size=14, color=NAVY)
        ws.merge_cells(
            start_row=r, start_column=c0,
            end_row=r, end_column=c0 + len(headers) - 1,
        )
        r += 1
        r += 1  # spacer

    # Header row
    header_row = r
    for i, h in enumerate(headers):
        cell = ws.cell(row=r, column=c0 + i, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', fgColor=NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _thin_border('bottom')
    ws.row_dimensions[r].height = 24
    r += 1

    # Data rows
    for row_idx, row in enumerate(rows):
        for i, val in enumerate(row):
            cell = ws.cell(row=r, column=c0 + i, value=val)
            cell.font = Font(size=11)
            cell.alignment = Alignment(
                horizontal='left' if i == 0 else 'center',
                vertical='center',
            )
            if alt_rows and row_idx % 2 == 1:
                cell.fill = PatternFill('solid', fgColor=GREY_LIGHT)
            if number_formats and headers[i] in number_formats:
                cell.number_format = number_formats[headers[i]]
        r += 1

    last_row = r - 1
    last_col = c0 + len(headers) - 1

    # Bottom border on last data row
    for i in range(len(headers)):
        ws.cell(row=last_row, column=c0 + i).border = _thin_border('bottom')

    # Freeze header
    if freeze_header:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=c0)

    # Auto-width
    if autofit:
        for i, h in enumerate(headers):
            col_letter = get_column_letter(c0 + i)
            width = max(len(str(h)) + 4, 10)
            for row in rows:
                if i < len(row):
                    width = max(width, min(60, len(str(row[i])) + 2))
            if column_widths and h in column_widths:
                width = column_widths[h]
            ws.column_dimensions[col_letter].width = width

    return header_row, last_row, last_col


def _thin_border(*sides):
    s = Side(style='thin', color='000000')
    kw = {side: s for side in sides}
    return Border(**kw)


# ─── Data validation (Likert dropdowns) ──────────────────────────────────

def add_likert_validation(ws, col_letter, first_row, last_row, scale='1-5'):
    if scale == '1-5':
        formula = '"1,2,3,4,5"'
    elif scale == '1-7':
        formula = '"1,2,3,4,5,6,7"'
    elif scale == '0-3':
        formula = '"0,1,2,3"'
    else:
        formula = f'"{scale}"'
    dv = DataValidation(type='list', formula1=formula, allow_blank=True)
    dv.error = 'Fuera de escala'
    dv.prompt = 'Seleccione una opción de la escala Likert'
    ws.add_data_validation(dv)
    dv.add(f'{col_letter}{first_row}:{col_letter}{last_row}')


def add_color_scale(ws, cell_range, kind='red_yellow_green'):
    if kind == 'red_yellow_green':
        rule = ColorScaleRule(
            start_type='min', start_color=BAD,
            mid_type='percentile', mid_value=50, mid_color=WARN,
            end_type='max', end_color=GOOD,
        )
    else:
        rule = ColorScaleRule(
            start_type='min', start_color='FFFFFF',
            end_type='max', end_color=NAVY,
        )
    ws.conditional_formatting.add(cell_range, rule)


# ─── Psychometric / statistical helpers ──────────────────────────────────

def cronbach_alpha(responses):
    """responses: list[list[number]]  (rows = subjects, cols = items).
    Returns (alpha, k, total_variance, sum_item_variances)."""
    if not responses or not responses[0]:
        return None, 0, 0.0, 0.0
    k = len(responses[0])
    # per-item variance across subjects
    item_variances = []
    for j in range(k):
        col = [r[j] for r in responses if j < len(r)]
        item_variances.append(statistics.pvariance(col) if len(col) > 1 else 0.0)
    # total score per subject
    totals = [sum(r[:k]) for r in responses]
    total_var = statistics.pvariance(totals) if len(totals) > 1 else 0.0
    if total_var == 0:
        return 0.0, k, 0.0, sum(item_variances)
    alpha = (k / (k - 1)) * (1 - sum(item_variances) / total_var)
    return alpha, k, total_var, sum(item_variances)


def _rank(values):
    """Tied-average ranking."""
    idx = sorted(range(len(values)), key=lambda i: values[i])
    ranks = [0] * len(values)
    i = 0
    while i < len(values):
        j = i
        while j + 1 < len(values) and values[idx[j + 1]] == values[idx[i]]:
            j += 1
        avg = (i + j) / 2 + 1  # 1-based rank
        for k in range(i, j + 1):
            ranks[idx[k]] = avg
        i = j + 1
    return ranks


def spearman(x, y):
    """Spearman rank correlation (no ties handling optimisation)."""
    rx, ry = _rank(x), _rank(y)
    mean_rx = sum(rx) / len(rx)
    mean_ry = sum(ry) / len(ry)
    num = sum((a - mean_rx) * (b - mean_ry) for a, b in zip(rx, ry))
    den_x = math.sqrt(sum((a - mean_rx) ** 2 for a in rx))
    den_y = math.sqrt(sum((b - mean_ry) ** 2 for b in ry))
    if den_x == 0 or den_y == 0:
        return 0.0
    return num / (den_x * den_y)


def spearman_matrix(df_like):
    """df_like: dict[column_name] -> list[number]. Returns (names, matrix)."""
    names = list(df_like.keys())
    n = len(names)
    matrix = [[1.0 if i == j else spearman(df_like[names[i]], df_like[names[j]])
               for j in range(n)] for i in range(n)]
    return names, matrix


def descriptives(values, label=''):
    vals = [v for v in values if isinstance(v, (int, float))]
    if not vals:
        return {'label': label, 'n': 0, 'mean': 0, 'sd': 0, 'median': 0, 'min': 0, 'max': 0}
    return {
        'label': label,
        'n': len(vals),
        'mean': statistics.mean(vals),
        'sd': statistics.pstdev(vals) if len(vals) > 1 else 0.0,
        'median': statistics.median(vals),
        'min': min(vals),
        'max': max(vals),
    }


# ─── One-shot builders ───────────────────────────────────────────────────

def build_likert_db(wb, *, sheet_name, headers, likert_cols=None, n_rows=30, likert_scale='1-5'):
    """Pre-wire a Likert response database.

    headers: list[str] for ALL columns (id, demo, items 1..k).
    likert_cols: list[str] for the columns that are Likert (get validation).
    n_rows: how many empty subject rows to pre-allocate.
    """
    ws = wb.active if len(wb.sheetnames) == 1 and wb.active.max_row <= 1 else wb.create_sheet(sheet_name)
    ws.title = sheet_name
    # header + empty rows
    write_table(ws, headers=headers, rows=[[None] * len(headers)] * n_rows, alt_rows=False)
    # data validation on the Likert columns
    if likert_cols:
        for name in likert_cols:
            if name in headers:
                col_letter = get_column_letter(headers.index(name) + 1)
                add_likert_validation(ws, col_letter, 2, n_rows + 1, likert_scale)
    return ws


def build_cronbach_sheet(wb, *, sheet_name, responses, label='Escala'):
    ws = wb.create_sheet(sheet_name)
    ws['A1'] = f'Fiabilidad · {label}'
    ws['A1'].font = Font(bold=True, size=16, color=NAVY)
    ws['A3'] = 'Resumen'
    ws['A3'].font = Font(bold=True, size=12)
    alpha, k, tv, siv = cronbach_alpha(responses)
    rows = [
        ['Ítems (k)', k],
        ['Sujetos (n)', len(responses)],
        ['Σ varianzas ítem', round(siv, 3)],
        ['Varianza total', round(tv, 3)],
        ["Cronbach's α", round(alpha, 3) if alpha is not None else 'n/a'],
        ['Interpretación', _alpha_label(alpha)],
    ]
    write_table(ws, headers=['Métrica', 'Valor'], rows=rows, start_row=4, freeze_header=False, alt_rows=True)
    return ws


def _alpha_label(a):
    if a is None: return 'n/a'
    if a >= 0.9: return 'Excelente'
    if a >= 0.8: return 'Bueno'
    if a >= 0.7: return 'Aceptable'
    if a >= 0.6: return 'Cuestionable'
    if a >= 0.5: return 'Pobre'
    return 'Inaceptable'


def build_spearman_sheet(wb, *, sheet_name, df_like):
    ws = wb.create_sheet(sheet_name)
    ws['A1'] = 'Correlaciones · Spearman ρ'
    ws['A1'].font = Font(bold=True, size=16, color=NAVY)
    names, mat = spearman_matrix(df_like)
    headers = [''] + names
    rows = [[names[i]] + [round(mat[i][j], 3) for j in range(len(names))] for i in range(len(names))]
    header_row, last_row, last_col = write_table(
        ws, headers=headers, rows=rows, start_row=3, alt_rows=False,
    )
    # Colour-scale on the numeric region only (skip first column of labels).
    from openpyxl.utils import get_column_letter as gcl
    first_data_col = gcl(2)
    last_data_col = gcl(last_col)
    add_color_scale(ws, f'{first_data_col}{header_row+1}:{last_data_col}{last_row}')
    return ws


def add_bar_chart(ws, *, title, data_range, categories_range, anchor_cell='H2'):
    chart = BarChart()
    chart.type = 'col'
    chart.style = 11
    chart.title = title
    chart.y_axis.title = None
    chart.x_axis.title = None
    data = Reference(ws, range_string=data_range)
    cats = Reference(ws, range_string=categories_range)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor_cell)
